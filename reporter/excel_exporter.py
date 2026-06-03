from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from reporter.judgment_labels import to_display_judgment


def save_listing_table(rows: Iterable[Mapping], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = list(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "\u641c\u7d22\u7ed3\u679c"

    headers = ["\u5e8f\u53f7", "\u6807\u9898", "\u4ef7\u683c", "\u5546\u54c1\u94fe\u63a5", "\u5224\u5b9a\u7ed3\u679c", "\u89c4\u683c\u91c7\u96c6\u6a21\u5f0f", "\u89c4\u683c\u91c7\u96c6\u4fe1\u606f"]
    ws.append(headers)

    sorted_rows = sorted(
        rows,
        key=lambda row: (float(row.get("price") or 0) <= 0, float(row.get("price") or 0)),
    )
    if sorted_rows:
        for index, row in enumerate(sorted_rows, start=1):
            title = str(row.get("title", ""))
            values = [
                index,
                title,
                row.get("price", ""),
                row.get("url", ""),
                to_display_judgment(str(row.get("judgment", ""))),
                row.get("spec_capture_mode", ""),
                row.get("spec_capture_info", ""),
            ]
            ws.append(values)
    else:
        values = [
            1,
            "\u672c\u6b21\u641c\u7d22\u672a\u91c7\u96c6\u5230\u7b26\u5408\u6807\u9898\u89c4\u5219\u7684\u5546\u54c1",
            "",
            "",
            "",
            "",
            "",
        ]
        ws.append(values)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 8,
        "B": 80,
        "C": 12,
        "D": 64,
        "E": 16,
        "F": 20,
        "G": 80,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows(min_row=2):
        row[0].alignment = Alignment(horizontal="center", vertical="top")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[2].alignment = Alignment(horizontal="right", vertical="top")
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        row[4].alignment = Alignment(horizontal="center", vertical="top")
        row[5].alignment = Alignment(horizontal="center", vertical="top")
        row[6].alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if ws.max_row >= 2:
        table_ref = f"A1:G{ws.max_row}"
        table = Table(displayName="SearchResults", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    wb.save(path)
    return path
